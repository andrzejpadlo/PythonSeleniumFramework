import openpyxl


class HomePageData:

    test_HomePage_data = [
        {"firstname": "Test", "email": "test@thissheet.com", "gender": "Female"},
        {"firstname": "Zbigniew", "email": "zstonoga@sejm.gov.pl", "gender": "Male"}
    ]

    @staticmethod
    def getTestData(test_case_name):
        book = openpyxl.load_workbook(
            "D:\\Python\\Selenium\\Udemy_Course_2\\PythonSelFramework\\TestData\\PythonDemo.xlsx")
        sheet = book.active
        Dict = {}
        for i in range(2, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return [Dict]
