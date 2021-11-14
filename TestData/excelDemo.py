import openpyxl

book = openpyxl.load_workbook("D:\\Python\\Selenium\\Udemy_Course_2\\PythonSelFramework\\TestData\\PythonDemo.xlsx")
sheet = book.active

# Read value from cell 1,2
print(sheet.cell(row=1, column=2).value)
print(sheet["B1"].value)

# Write value into cell 2,2
sheet.cell(row=2, column=2).value = "Taste"

print(sheet.cell(row=2, column=2).value)

# Get the number of rows and columns
print(sheet.max_row)
print(sheet.max_column)

# Print all available cells
for i in range(1, sheet.max_row+1):
    for j in range(1, sheet.max_column+1):
        print(sheet.cell(row=i, column=j).value)

# Put excel data into dictionary
Dict = {}
for i in range(2, sheet.max_row+1):
    for j in range(2, sheet.max_column+1):
        Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
print(Dict)
